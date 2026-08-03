from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from tkinter import BooleanVar, StringVar, Tk, filedialog, messagebox
from tkinter import ttk

try:
    from PIL import Image, ImageTk
except Exception:  # pragma: no cover - handled in GUI startup
    Image = None
    ImageTk = None


APP_NAME = "NamecardUpdater"
PAGES_URL = "https://jianshenghao2023-creator.github.io/Namecard/"

CONTACT_COLUMNS = [
    "record_id",
    "review_status",
    "confidence",
    "source_pdf",
    "source_page",
    "card_position",
    "source_image",
    "person_name",
    "chinese_name",
    "company",
    "company_normalized",
    "title",
    "industry_tags",
    "relationship_stage",
    "country",
    "city",
    "email_primary",
    "email_secondary",
    "mobile",
    "phone",
    "website",
    "address",
    "notes",
]

ENRICHED_APPEND_COLUMNS = [
    "primary_category_code",
    "primary_category",
    "subcategory",
    "business_role",
    "classification_confidence",
    "verification_source_url",
    "online_check_status",
    "verified_on",
    "classification_basis",
]

ENRICHED_COLUMNS = CONTACT_COLUMNS + ENRICHED_APPEND_COLUMNS

WEB_FIELDS = [
    "record_id",
    "review_status",
    "confidence",
    "source_pdf",
    "source_page",
    "card_position",
    "source_image",
    "person_name",
    "chinese_name",
    "company",
    "company_normalized",
    "title",
    "industry_tags",
    "relationship_stage",
    "country",
    "city",
    "email_primary",
    "email_secondary",
    "mobile",
    "phone",
    "website",
    "address",
    "notes",
    "primary_category_code",
    "primary_category",
    "subcategory",
    "business_role",
    "classification_confidence",
    "online_check_status",
    "verified_on",
    "classification_basis",
]

COMPANY_COLUMNS = [
    "company_normalized",
    "primary_category_code",
    "primary_category",
    "subcategory",
    "business_role",
    "classification_confidence",
    "contact_count",
    "country_samples",
    "website_samples",
    "industry_tags_from_cards",
    "source_pages",
    "classification_basis",
    "review_status",
    "notes",
]

VERIFIED_COMPANY_COLUMNS = [
    "company_normalized",
    "primary_category_code",
    "primary_category",
    "subcategory",
    "business_role",
    "classification_confidence",
    "contact_count",
    "country_samples",
    "website_samples",
    "industry_tags_from_cards",
    "source_pages",
    "verification_source_url",
    "online_check_status",
    "verified_on",
    "classification_basis",
    "review_status",
    "notes",
]

CATEGORIES = {
    "C01": ("船东/船管/航运运营", "拥有、管理、运营船队，或作为班轮/散货/专线航运公司使用船舶资产。"),
    "C02": ("船厂/修造/船舶工程", "造船、修船、改装、船舶工程和船厂体系内业务单位。"),
    "C03": ("船舶经纪/租船买卖中介", "船舶买卖、新造船、租船、S&P 经纪与撮合机构。"),
    "C04": ("港口/码头/物流/供应链", "港口、码头、港口物流、货代、供应链和区域代表机构。"),
    "C05": ("主机/动力/推进/传动核心设备", "主机、发电机组、推进器、齿轮箱、涡轮增压、燃烧系统、船舶自动化等核心设备。"),
    "C06": ("LNG/低温/燃气/能源系统", "LNG、低温、燃气、工业气体、燃料泵、气体工程和相关能源系统。"),
    "C07": ("船舶/船厂通用设备、材料、安全与内装供应商", "泵阀、换热器、救生消防、甲板机械、内装材料、节能设备、特殊材料、船厂生产装备等。"),
    "C08": ("船级社/研发试验/技术服务", "船级社、检验认证、船模试验、水动力研发、工程咨询、船舶技术服务。"),
    "C09": ("航运金融/能源货主/其他支持", "航运金融、能源公司/货主、研究合作方、差旅等外围但可能有业务价值的对象。"),
}

COMPANY_TERMS = [
    "gmbh",
    "kg",
    "bv",
    "b.v.",
    "ag",
    "a/s",
    "oy",
    "s.a.",
    "s.a",
    "sas",
    "srl",
    "ltd",
    "limited",
    "llc",
    "inc",
    "corp",
    "corporation",
    "marine",
    "maritime",
    "shipping",
    "ship",
    "reederei",
    "designers",
    "constructors",
    "systems",
    "solutions",
    "group",
    "technologies",
    "technology",
    "engineering",
]

TITLE_TERMS = [
    "manager",
    "director",
    "engineer",
    "captain",
    "sales",
    "project",
    "president",
    "ceo",
    "coo",
    "cto",
    "head",
    "superintendent",
    "broker",
    "consultant",
    "coordinator",
    "managing",
    "business",
]

EMAIL_RE = re.compile(r"[\w.+\-]+@[\w.\-]+\.[A-Za-z]{2,}", re.I)
URL_RE = re.compile(r"((?:https?://)?(?:www\.)?[A-Za-z0-9][A-Za-z0-9.\-]+\.[A-Za-z]{2,}(?:/[^\s]*)?)", re.I)
PHONE_RE = re.compile(r"(?:(?:phone|mobile|mob|tel|fax|direct)\s*)?(\+?\d[\d\s()./\-]{5,}\d)", re.I)


@dataclass
class ContactDraft:
    data: dict[str, str] = field(default_factory=dict)
    ocr_text: str = ""
    crop_image: str = ""

    def get(self, key: str) -> str:
        return str(self.data.get(key, "") or "")

    def set(self, key: str, value: str) -> None:
        self.data[key] = value or ""


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def candidate_project_roots() -> list[Path]:
    candidates: list[Path] = []
    for start in [Path.cwd(), app_dir(), Path("E:/namecard")]:
        try:
            start = start.resolve()
        except OSError:
            continue
        candidates.append(start)
        candidates.extend(start.parents)
    out: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate).lower()
        if key not in seen:
            seen.add(key)
            out.append(candidate)
    return out


def find_project_root() -> Path:
    for root in candidate_project_roots():
        if (root / "mobile_search").is_dir() and (
            (root / "namecard_contacts_enriched_v1.csv").is_file()
            or (root / "namecard_contacts_template_v1.csv").is_file()
        ):
            return root
    return Path("E:/namecard")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def write_csv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def select_columns(row: dict[str, object], columns: list[str]) -> dict[str, str]:
    return {column: str(row.get(column, "") or "") for column in columns}


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_email(text: str) -> str:
    return (text or "").strip().strip(".,;:").lower()


def source_basename(row: dict[str, str]) -> str:
    raw = row.get("source_pdf", "")
    return Path(raw).name if raw else ""


def source_page_key(row: dict[str, str]) -> str:
    name = source_basename(row) or row.get("source_pdf", "")
    return f"{name}:{row.get('source_page', '')}"


def split_tags(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[;；]", value or "") if part.strip()]


def join_unique(values: list[str], separator: str = "; ") -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        value = clean_text(value)
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return separator.join(out)


def normalize_lookup_text(value: str) -> str:
    value = clean_text(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().replace("&", "and")
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", value)


def contact_display_name(row: dict[str, str]) -> str:
    return clean_text(row.get("person_name", "") or row.get("chinese_name", ""))


def contact_company(row: dict[str, str]) -> str:
    return clean_text(row.get("company_normalized", "") or row.get("company", ""))


def contact_label(row: dict[str, str], fallback: str = "") -> str:
    record_id = clean_text(row.get("record_id", ""))
    name = contact_display_name(row)
    company = contact_company(row)
    email = clean_text(row.get("email_primary", ""))
    parts = [part for part in [record_id, name, company, email] if part]
    return " | ".join(parts) if parts else fallback


def contact_emails(row: dict[str, str]) -> set[str]:
    emails: set[str] = set()
    for field_name in ["email_primary", "email_secondary"]:
        raw = row.get(field_name, "") or ""
        for match in EMAIL_RE.finditer(raw):
            emails.add(normalize_email(match.group(0)))
    return emails


def phone_key(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if digits.startswith("00"):
        digits = digits[2:]
    if len(digits) < 7:
        return ""
    return digits[-9:] if len(digits) >= 9 else digits


def contact_phone_keys(row: dict[str, str]) -> set[str]:
    keys: set[str] = set()
    for field_name in ["mobile", "phone"]:
        raw = row.get(field_name, "") or ""
        for match in PHONE_RE.finditer(raw):
            key = phone_key(match.group(1))
            if key:
                keys.add(key)
        key = phone_key(raw)
        if key:
            keys.add(key)
    return keys


def contact_name_company_key(row: dict[str, str]) -> str:
    name = normalize_lookup_text(contact_display_name(row))
    company = normalize_lookup_text(contact_company(row))
    if not name or not company:
        return ""
    return f"{name}|{company}"


def find_duplicate_reports(
    drafts: list[ContactDraft],
    existing_rows: list[dict[str, str]],
    replace_same_pdf: bool,
) -> dict[int, list[str]]:
    source_names = {source_basename(draft.data).lower() for draft in drafts if source_basename(draft.data)}
    filtered_existing = []
    for row in existing_rows:
        if replace_same_pdf and source_basename(row).lower() in source_names:
            continue
        filtered_existing.append(row)

    email_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    phone_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    name_company_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in filtered_existing:
        for email in contact_emails(row):
            email_index[email].append(row)
        for key in contact_phone_keys(row):
            phone_index[key].append(row)
        key = contact_name_company_key(row)
        if key:
            name_company_index[key].append(row)

    reports: dict[int, list[str]] = defaultdict(list)
    for index, draft in enumerate(drafts):
        row = select_columns(draft.data, ENRICHED_COLUMNS)
        for email in sorted(contact_emails(row)):
            for match in email_index.get(email, []):
                reports[index].append(f"邮箱重复：{email} -> {contact_label(match)}")
        for key in sorted(contact_phone_keys(row)):
            for match in phone_index.get(key, []):
                reports[index].append(f"电话重复：尾号/号码 {key} -> {contact_label(match)}")
        key = contact_name_company_key(row)
        for match in name_company_index.get(key, []):
            reports[index].append(f"姓名+公司重复 -> {contact_label(match)}")

    batch_rows = [select_columns(draft.data, ENRICHED_COLUMNS) for draft in drafts]
    for index, row in enumerate(batch_rows):
        row_emails = contact_emails(row)
        row_phones = contact_phone_keys(row)
        row_name_company = contact_name_company_key(row)
        for other_index, other in enumerate(batch_rows):
            if other_index <= index:
                continue
            label = contact_label(other, fallback=f"第 {other_index + 1} 条")
            email_overlap = row_emails & contact_emails(other)
            phone_overlap = row_phones & contact_phone_keys(other)
            if email_overlap:
                reports[index].append(f"本批邮箱重复：{', '.join(sorted(email_overlap))} -> {label}")
                reports[other_index].append(f"本批邮箱重复：{', '.join(sorted(email_overlap))} -> {contact_label(row, fallback=f'第 {index + 1} 条')}")
            if phone_overlap:
                reports[index].append(f"本批电话重复：{', '.join(sorted(phone_overlap))} -> {label}")
                reports[other_index].append(f"本批电话重复：{', '.join(sorted(phone_overlap))} -> {contact_label(row, fallback=f'第 {index + 1} 条')}")
            if row_name_company and row_name_company == contact_name_company_key(other):
                reports[index].append(f"本批姓名+公司重复 -> {label}")
                reports[other_index].append(f"本批姓名+公司重复 -> {contact_label(row, fallback=f'第 {index + 1} 条')}")

    return {index: join_unique(items, "\n").split("\n") for index, items in reports.items() if items}


def duplicate_summary(items: list[str]) -> str:
    if not items:
        return ""
    labels: list[str] = []
    for item in items:
        labels.append(item.split("：", 1)[0].split(" -> ", 1)[0])
    return join_unique(labels, "/") or "可能重复"


def ensure_url(raw: str) -> str:
    raw = clean_text(raw)
    if not raw:
        return ""
    first = raw.split(";")[0].strip()
    if not first:
        return ""
    if not first.startswith(("http://", "https://")):
        first = "https://" + first
    return first


def company_from_email_domain(email: str, domain_map: dict[str, str]) -> str:
    if "@" not in email:
        return ""
    domain = email.split("@", 1)[1].lower()
    parts = domain.split(".")
    for i in range(len(parts) - 1):
        candidate = ".".join(parts[i:])
        if candidate in domain_map:
            return domain_map[candidate]
    return ""


def build_domain_map(rows: list[dict[str, str]]) -> dict[str, str]:
    domain_map: dict[str, str] = {}
    for row in rows:
        company = row.get("company_normalized") or row.get("company") or ""
        if not company:
            continue
        for field_name in ["email_primary", "email_secondary"]:
            email = normalize_email(row.get(field_name, ""))
            if "@" in email:
                domain = email.split("@", 1)[1]
                domain_map.setdefault(domain, company)
        website = ensure_url(row.get("website", ""))
        if website:
            host = urllib.parse.urlparse(website).netloc.lower().removeprefix("www.")
            if host:
                domain_map.setdefault(host, company)
    return domain_map


def looks_like_company(line: str) -> bool:
    lower = line.lower()
    return any(term in lower for term in COMPANY_TERMS)


def looks_like_title(line: str) -> bool:
    lower = line.lower()
    return any(term in lower for term in TITLE_TERMS)


def looks_like_name(line: str) -> bool:
    if not line or len(line) > 45:
        return False
    lower = line.lower()
    if any(token in lower for token in ["@", "www", "http", "phone", "mobile", "fax", "mail"]):
        return False
    if looks_like_company(line):
        return False
    words = [word for word in re.split(r"\s+", line.strip()) if word]
    if not (1 <= len(words) <= 5):
        return False
    letters = sum(ch.isalpha() for ch in line)
    return letters >= max(3, len(line) // 2)


def known_company_match(text: str, companies: list[str]) -> str:
    lower = text.lower()
    sorted_companies = sorted((c for c in companies if c), key=len, reverse=True)
    for company in sorted_companies:
        if len(company) >= 5 and company.lower() in lower:
            return company
    return ""


def infer_tags(company: str, title: str, ocr_text: str) -> str:
    text = f"{company} {title} {ocr_text}".lower()
    tags: list[str] = []
    rules = [
        ("shipowner shipping reederei fleet vessel containership container", "船东/船管;航运"),
        ("shipbroker broker chartering sale purchase newbuilding", "船舶经纪;租船买卖"),
        ("port terminal logistics supply chain", "港口/码头;物流"),
        ("shipyard shipbuilding repair conversion", "船厂;修造船"),
        ("engine propulsion thruster gearbox turbocharger navigation automation radar", "主机/动力/推进/传动;核心设备"),
        ("lng cryogenic gas fuel hydrogen", "LNG/低温/燃气;能源系统"),
        ("pump valve hatch cover roro ro-ro crane winch fire safety lifeboat equipment system", "船舶设备;设备供应商"),
        ("class classification survey certification research model basin", "船级社/认证;技术服务"),
    ]
    for words, tag in rules:
        if any(word in text for word in words.split()):
            tags.extend(split_tags(tag))
    return join_unique(tags) or "待确认"


def infer_classification(company: str, tags: str, title: str, website: str, ocr_text: str) -> dict[str, str]:
    text = f"{company} {tags} {title} {website} {ocr_text}".lower()
    checks = [
        ("C01", ["船东", "船管", "航运", "shipowner", "shipping", "reederei", "fleet", "containership", "vessel"], "船东/船管/航运运营", "潜在客户/船舶运营方"),
        ("C02", ["船厂", "修造", "shipyard", "shipbuilding", "repair", "conversion"], "船厂/修造船", "船厂/项目方"),
        ("C03", ["经纪", "broker", "shipbroker", "chartering", "sale & purchase", "s&p"], "船舶经纪/租船买卖", "中间商/经纪"),
        ("C04", ["港口", "码头", "物流", "port", "terminal", "logistics", "supply chain"], "港口/码头/物流", "港口物流服务商"),
        ("C05", ["主机", "推进", "齿轮", "雷达", "导航", "engine", "propulsion", "thruster", "gear", "automation", "radar"], "主机/动力/推进/传动核心设备", "核心设备供应商"),
        ("C06", ["lng", "低温", "燃气", "cryogenic", "gas", "fuel", "hydrogen"], "LNG/低温/燃气/能源系统", "燃气与低温系统供应商"),
        ("C07", ["设备", "供应商", "pump", "valve", "hatch", "cover", "roro", "ro-ro", "crane", "winch", "safety", "fire", "lifeboat", "equipment"], "船舶通用设备/材料/安全/内装", "设备供应商"),
        ("C08", ["船级", "认证", "研发", "classification", "class", "survey", "certification", "research"], "船级社/研发试验/技术服务", "认证/技术服务"),
    ]
    for code, needles, subcategory, role in checks:
        if any(needle in text for needle in needles):
            return {
                "primary_category_code": code,
                "primary_category": CATEGORIES[code][0],
                "subcategory": subcategory,
                "business_role": role,
                "classification_confidence": "medium",
                "review_status": "待确认",
            }
    return {
        "primary_category_code": "C09",
        "primary_category": CATEGORIES["C09"][0],
        "subcategory": "待确认",
        "business_role": "待确认",
        "classification_confidence": "low",
        "review_status": "待确认",
    }


def fetch_website_hint(url: str) -> tuple[str, str]:
    url = ensure_url(url)
    if not url:
        return "no_url", ""
    try:
        request = urllib.request.Request(url, headers={"User-Agent": f"{APP_NAME}/1.0"})
        with urllib.request.urlopen(request, timeout=8) as response:
            status = f"http_{getattr(response, 'status', 200)}"
            raw = response.read(200_000)
        text = raw.decode("utf-8", errors="ignore")
        title = re.search(r"<title[^>]*>(.*?)</title>", text, flags=re.I | re.S)
        desc = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\'](.*?)["\']', text, flags=re.I | re.S)
        parts = []
        if title:
            parts.append(clean_text(re.sub(r"<[^>]+>", " ", title.group(1))))
        if desc:
            parts.append(clean_text(re.sub(r"<[^>]+>", " ", desc.group(1))))
        return status, " | ".join(parts)[:500]
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        return f"request_failed:{exc.__class__.__name__}", ""


def find_tesseract() -> str:
    env_cmd = os.environ.get("TESSERACT_CMD", "")
    candidates = [
        env_cmd,
        shutil.which("tesseract") or "",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        str(app_dir() / "tesseract" / "tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return candidate
    return ""


def tesseract_languages(tesseract_cmd: str) -> list[str]:
    if not tesseract_cmd:
        return []
    try:
        result = subprocess.run([tesseract_cmd, "--list-langs"], capture_output=True, text=True, timeout=10)
    except (OSError, subprocess.SubprocessError):
        return []
    lines = [line.strip() for line in (result.stdout + "\n" + result.stderr).splitlines()]
    return [line for line in lines if line and not line.lower().startswith("list of")]


def choose_ocr_language(tesseract_cmd: str) -> str:
    langs = set(tesseract_languages(tesseract_cmd))
    preferred = [lang for lang in ["eng", "deu", "fra", "chi_sim", "chi_tra"] if lang in langs]
    return "+".join(preferred or ["eng"])


def run_ocr(tesseract_cmd: str, image_path: Path, language: str) -> str:
    if not tesseract_cmd:
        return ""
    try:
        result = subprocess.run(
            [tesseract_cmd, str(image_path), "stdout", "-l", language or "eng", "--psm", "6"],
            capture_output=True,
            text=True,
            timeout=90,
        )
    except (OSError, subprocess.SubprocessError) as exc:
        return f"[OCR failed: {exc}]"
    if result.returncode != 0:
        return f"[OCR failed: {result.stderr.strip()}]"
    return result.stdout.strip()


def render_pdf(pdf_path: Path, output_dir: Path, scale: float = 3.0) -> list[Path]:
    try:
        import pypdfium2 as pdfium
    except Exception as exc:
        raise RuntimeError("缺少 pypdfium2，无法渲染 PDF。请重新构建 EXE 或安装依赖。") from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    document = pdfium.PdfDocument(str(pdf_path))
    pages: list[Path] = []
    for index in range(len(document)):
        page = document[index]
        bitmap = page.render(scale=scale)
        image = bitmap.to_pil()
        path = output_dir / f"page_{index + 1:02d}.jpg"
        image.save(path, quality=92)
        pages.append(path)
    return pages


def intervals_from_projection(values: list[int], threshold: int, min_len: int, merge_gap: int) -> list[tuple[int, int]]:
    intervals: list[tuple[int, int]] = []
    start: int | None = None
    for index, value in enumerate(values):
        if value > threshold and start is None:
            start = index
        elif value <= threshold and start is not None:
            if index - start >= min_len:
                intervals.append((start, index))
            start = None
    if start is not None and len(values) - start >= min_len:
        intervals.append((start, len(values)))
    if not intervals:
        return []
    merged = [intervals[0]]
    for start, end in intervals[1:]:
        last_start, last_end = merged[-1]
        if start - last_end <= merge_gap:
            merged[-1] = (last_start, end)
        else:
            merged.append((start, end))
    return merged


def content_bbox(image) -> tuple[int, int, int, int]:
    width, height = image.size
    pixels = image.load()
    xs: list[int] = []
    ys: list[int] = []
    for y in range(int(height * 0.01), int(height * 0.95), 4):
        for x in range(int(width * 0.03), int(width * 0.97), 4):
            if pixels[x, y] < 235:
                xs.append(x)
                ys.append(y)
    if not xs or not ys:
        return (0, 0, width, height)
    return (max(0, min(xs) - 40), max(0, min(ys) - 40), min(width, max(xs) + 40), min(height, max(ys) + 40))


def grid_card_crops(page_image: Path, output_dir: Path) -> list[tuple[str, Path]]:
    if Image is None:
        return []
    gray = Image.open(page_image).convert("L")
    rgb = Image.open(page_image).convert("RGB")
    width, height = gray.size
    x0, y0, x1, y1 = content_bbox(gray)
    content_width = x1 - x0
    content_height = y1 - y0
    if content_width < width * 0.2 or content_height < height * 0.08:
        return []

    pixels = gray.load()
    left_dark = 0
    right_dark = 0
    for y in range(y0, y1, 5):
        for x in range(x0, x0 + content_width // 2, 5):
            if pixels[x, y] < 235:
                left_dark += 1
        for x in range(x0 + content_width // 2, x1, 5):
            if pixels[x, y] < 235:
                right_dark += 1
    columns = 2 if content_width > width * 0.55 and left_dark > 50 and right_dark > 50 else 1
    target_row_height = max(260, (content_width / columns) * 0.72)
    rows = max(1, min(6, round(content_height / target_row_height)))

    crops: list[tuple[str, Path]] = []
    card_index = 1
    for row in range(rows):
        row_start = int(y0 + row * content_height / rows)
        row_end = int(y0 + (row + 1) * content_height / rows)
        for column in range(columns):
            col_start = int(x0 + column * content_width / columns)
            col_end = int(x0 + (column + 1) * content_width / columns)
            crop = rgb.crop((max(0, col_start - 20), max(0, row_start - 25), min(width, col_end + 20), min(height, row_end + 25)))
            crop_path = output_dir / f"{page_image.stem}_grid_card_{card_index:02d}.jpg"
            crop.save(crop_path, quality=95)
            crops.append((f"第{card_index}张", crop_path))
            card_index += 1
    return crops


def detect_card_crops(page_image: Path, output_dir: Path) -> list[tuple[str, Path]]:
    if Image is None:
        return []
    output_dir.mkdir(parents=True, exist_ok=True)
    image = Image.open(page_image).convert("L")
    width, height = image.size
    pixels = image.load()
    row_values: list[int] = []
    for y in range(height):
        count = 0
        for x in range(0, width, 4):
            if pixels[x, y] < 238:
                count += 1
        row_values.append(count)
    row_intervals = intervals_from_projection(row_values, threshold=max(4, width // 900), min_len=max(80, height // 40), merge_gap=max(70, height // 35))
    if not row_intervals:
        row_intervals = [(0, height)]

    rgb = Image.open(page_image).convert("RGB")
    crops: list[tuple[str, Path]] = []
    card_index = 1
    for row_start, row_end in row_intervals:
        col_values: list[int] = []
        for x in range(width):
            count = 0
            for y in range(row_start, row_end, 4):
                if pixels[x, y] < 238:
                    count += 1
            col_values.append(count)
        col_intervals = intervals_from_projection(col_values, threshold=max(3, (row_end - row_start) // 900), min_len=max(180, width // 8), merge_gap=max(70, width // 30))
        if len(col_intervals) == 1 and (col_intervals[0][1] - col_intervals[0][0]) > width * 0.68:
            left_dark = sum(col_values[int(width * 0.08) : int(width * 0.42)])
            right_dark = sum(col_values[int(width * 0.58) : int(width * 0.92)])
            if left_dark > 50 and right_dark > 50:
                search_start = int(width * 0.38)
                search_end = int(width * 0.62)
                window = max(20, width // 80)
                best_x = width // 2
                best_value = None
                for x in range(search_start, search_end):
                    value = sum(col_values[max(0, x - window) : min(width, x + window)])
                    if best_value is None or value < best_value:
                        best_value = value
                        best_x = x
                left = (col_intervals[0][0], best_x)
                right = (best_x, col_intervals[0][1])
                if left[1] - left[0] > width * 0.2 and right[1] - right[0] > width * 0.2:
                    col_intervals = [left, right]
        if not col_intervals:
            col_intervals = [(0, width)]
        for col_start, col_end in col_intervals:
            x0 = max(0, col_start - 50)
            y0 = max(0, row_start - 50)
            x1 = min(width, col_end + 50)
            y1 = min(height, row_end + 50)
            if x1 - x0 < width * 0.18 or y1 - y0 < height * 0.05:
                continue
            crop = rgb.crop((x0, y0, x1, y1))
            crop_path = output_dir / f"{page_image.stem}_card_{card_index:02d}.jpg"
            crop.save(crop_path, quality=95)
            crops.append((f"第{card_index}张", crop_path))
            card_index += 1
    grid_crops = grid_card_crops(page_image, output_dir)
    if len(crops) < 4 and len(grid_crops) > len(crops):
        return grid_crops
    if not crops:
        crops.append(("整页", page_image))
    return crops


def parse_contact_from_ocr(
    text: str,
    source_pdf: Path,
    source_page: int,
    card_position: str,
    crop_path: Path,
    existing_rows: list[dict[str, str]],
) -> ContactDraft:
    lines = [clean_text(line) for line in text.splitlines()]
    lines = [line for line in lines if line and not line.startswith("[OCR failed")]
    full_text = "\n".join(lines)
    emails = [normalize_email(match.group(0)) for match in EMAIL_RE.finditer(full_text)]
    email = emails[0] if emails else ""

    websites = []
    for match in URL_RE.finditer(full_text):
        candidate = match.group(1).strip().strip(".,;:")
        if "@" not in candidate and not candidate.lower().endswith((".de", ".com", ".net", ".nl", ".gr", ".cn", ".fr", ".no", ".dk", ".fi", ".se", ".uk")):
            continue
        if "@" not in candidate:
            websites.append(candidate)
    website = websites[0] if websites else ""

    phone = ""
    mobile = ""
    for line in lines:
        lower = line.lower()
        for match in PHONE_RE.finditer(line):
            number = clean_text(match.group(1))
            if len(re.sub(r"\D", "", number)) < 7:
                continue
            if "mobile" in lower or "mob" in lower:
                mobile = mobile or number
            elif "fax" not in lower:
                phone = phone or number
    if not mobile:
        numbers = [clean_text(match.group(1)) for match in PHONE_RE.finditer(full_text)]
        numbers = [num for num in numbers if len(re.sub(r"\D", "", num)) >= 7]
        if numbers:
            phone = phone or numbers[0]
        if len(numbers) > 1:
            mobile = numbers[1]

    domain_map = build_domain_map(existing_rows)
    known_companies = sorted({row.get("company_normalized") or row.get("company") or "" for row in existing_rows})
    company = company_from_email_domain(email, domain_map)
    if not company:
        company = known_company_match(full_text, known_companies)
    if not company:
        company_candidates = [line for line in lines if looks_like_company(line)]
        company = company_candidates[0] if company_candidates else ""

    title = ""
    for line in lines:
        if looks_like_title(line):
            title = line
            break

    person_name = ""
    for line in lines:
        if line == company or line == title:
            continue
        if looks_like_name(line):
            person_name = line
            break

    address_lines = []
    for line in lines:
        lower = line.lower()
        if line in {person_name, company, title}:
            continue
        if EMAIL_RE.search(line) or URL_RE.search(line) or "phone" in lower or "mobile" in lower or "fax" in lower or "mail" in lower:
            continue
        if re.search(r"\d{4,}", line) or any(token in lower for token in ["street", "str.", "strasse", "avenue", "road", "po box", "hamburg", "drachten"]):
            address_lines.append(line)

    tags = infer_tags(company, title, full_text)
    classification = infer_classification(company, tags, title, website, full_text)
    data = {column: "" for column in ENRICHED_COLUMNS}
    data.update(
        {
            "record_id": "",
            "review_status": "待校对",
            "confidence": "medium" if full_text else "low",
            "source_pdf": str(source_pdf),
            "source_page": str(source_page),
            "card_position": card_position,
            "source_image": str(crop_path),
            "person_name": person_name,
            "company": company,
            "company_normalized": company,
            "title": title,
            "industry_tags": tags,
            "relationship_stage": "名片已扫描",
            "email_primary": email,
            "mobile": mobile,
            "phone": phone,
            "website": website,
            "address": join_unique(address_lines),
            "notes": "",
            "primary_category_code": classification["primary_category_code"],
            "primary_category": classification["primary_category"],
            "subcategory": classification["subcategory"],
            "business_role": classification["business_role"],
            "classification_confidence": classification["classification_confidence"],
            "online_check_status": "not_checked",
            "verified_on": "",
            "classification_basis": f"免费OCR与规则推断：{tags}",
        }
    )
    return ContactDraft(data=data, ocr_text=text, crop_image=str(crop_path))


def write_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, object]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    for index, (sheet_name, columns, rows) in enumerate(sheets):
        ws = wb.active if index == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(columns)
        for row in rows:
            ws.append([row.get(column, "") for column in columns])
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9D9D9")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
    wb.save(path)


def load_template_contacts(project_root: Path) -> list[dict[str, str]]:
    template = project_root / "namecard_contacts_template_v1.csv"
    if template.is_file():
        return [select_columns(row, CONTACT_COLUMNS) for row in read_csv(template)]
    enriched = project_root / "namecard_contacts_enriched_v1.csv"
    return [select_columns(row, CONTACT_COLUMNS) for row in read_csv(enriched)]


def build_company_rows(
    all_contacts: list[dict[str, str]],
    old_class_rows: list[dict[str, str]],
    new_drafts: list[ContactDraft],
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    today = datetime.now().date().isoformat()
    class_by_company = {row.get("company_normalized", ""): dict(row) for row in old_class_rows if row.get("company_normalized")}

    for draft in new_drafts:
        company = draft.get("company_normalized") or draft.get("company")
        if not company:
            continue
        code = draft.get("primary_category_code")
        if not code:
            code = infer_classification(company, draft.get("industry_tags"), draft.get("title"), draft.get("website"), draft.ocr_text)["primary_category_code"]
        status, hint = fetch_website_hint(draft.get("website"))
        basis = draft.get("classification_basis")
        if hint:
            basis = f"官网/网页标题辅助核验：{hint}；名片标签：{draft.get('industry_tags')}"
        row = class_by_company.get(company, {"company_normalized": company})
        row.update(
            {
                "company_normalized": company,
                "primary_category_code": code,
                "primary_category": CATEGORIES.get(code, CATEGORIES["C09"])[0],
                "subcategory": draft.get("subcategory") or row.get("subcategory") or "待确认",
                "business_role": draft.get("business_role") or row.get("business_role") or "待确认",
                "classification_confidence": draft.get("classification_confidence") or row.get("classification_confidence") or "medium",
                "verification_source_url": ensure_url(draft.get("website")) or row.get("verification_source_url", ""),
                "online_check_status": status if draft.get("website") else row.get("online_check_status", "no_url"),
                "verified_on": today if draft.get("website") else row.get("verified_on", ""),
                "classification_basis": basis,
                "review_status": "待确认" if (draft.get("classification_confidence") or "medium") in {"medium", "low"} else "已核验",
                "notes": row.get("notes", ""),
            }
        )
        class_by_company[company] = row

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for contact in all_contacts:
        company = contact.get("company_normalized") or contact.get("company")
        if company:
            grouped[company].append(contact)

    company_rows: list[dict[str, object]] = []
    for company in sorted(grouped):
        class_row = class_by_company.get(company)
        if not class_row:
            sample = grouped[company][0]
            inferred = infer_classification(company, sample.get("industry_tags", ""), sample.get("title", ""), sample.get("website", ""), "")
            class_row = {
                "company_normalized": company,
                **inferred,
                "verification_source_url": ensure_url(sample.get("website", "")),
                "online_check_status": "not_checked",
                "verified_on": "",
                "classification_basis": f"自动规则初分：{sample.get('industry_tags', '')}",
                "notes": "",
            }
        code = class_row.get("primary_category_code") or "C09"
        contacts = grouped[company]
        tags: list[str] = []
        countries: list[str] = []
        websites: list[str] = []
        pages: list[str] = []
        for contact in contacts:
            tags.extend(split_tags(contact.get("industry_tags", "")))
            countries.append(contact.get("country", ""))
            websites.append(contact.get("website", ""))
            pages.append(source_page_key(contact))
        company_rows.append(
            {
                "company_normalized": company,
                "primary_category_code": code,
                "primary_category": CATEGORIES.get(code, CATEGORIES["C09"])[0],
                "subcategory": class_row.get("subcategory", ""),
                "business_role": class_row.get("business_role", ""),
                "classification_confidence": class_row.get("classification_confidence", ""),
                "contact_count": str(len(contacts)),
                "country_samples": join_unique(countries),
                "website_samples": join_unique(websites),
                "industry_tags_from_cards": join_unique(tags),
                "source_pages": join_unique(pages),
                "verification_source_url": class_row.get("verification_source_url", ""),
                "online_check_status": class_row.get("online_check_status", ""),
                "verified_on": class_row.get("verified_on", ""),
                "classification_basis": class_row.get("classification_basis", ""),
                "review_status": class_row.get("review_status", ""),
                "notes": class_row.get("notes", ""),
            }
        )

    summary_rows: list[dict[str, object]] = []
    for code, (category, definition) in CATEGORIES.items():
        rows = [row for row in company_rows if row["primary_category_code"] == code]
        contact_count = sum(int(row["contact_count"]) for row in rows)
        summary_rows.append(
            {
                "category_code": code,
                "primary_category": category,
                "company_count": len(rows),
                "contact_count": contact_count,
                "definition": definition,
            }
        )
    return company_rows, summary_rows


def regenerate_outputs(project_root: Path, new_drafts: list[ContactDraft], replace_same_pdf: bool) -> tuple[int, int, str]:
    if not new_drafts:
        raise RuntimeError("没有可写入的新联系人。")
    contacts = load_template_contacts(project_root)
    source_pdf_name = source_basename(new_drafts[0].data)
    if replace_same_pdf and source_pdf_name:
        contacts = [row for row in contacts if source_basename(row).lower() != source_pdf_name.lower()]

    new_contacts: list[dict[str, str]] = []
    for draft in new_drafts:
        data = select_columns(draft.data, CONTACT_COLUMNS)
        data["review_status"] = data.get("review_status") or "待校对"
        data["confidence"] = data.get("confidence") or "medium"
        data["relationship_stage"] = data.get("relationship_stage") or "名片已扫描"
        data["company_normalized"] = data.get("company_normalized") or data.get("company")
        if not data["person_name"] and not data["chinese_name"]:
            raise RuntimeError("存在未填写姓名的联系人，请先确认表格。")
        if not data["company_normalized"]:
            raise RuntimeError(f"{data.get('person_name')} 未填写公司。")
        new_contacts.append(data)

    all_contacts = contacts + new_contacts
    for index, row in enumerate(all_contacts, start=1):
        row["record_id"] = f"NC-{index:04d}"

    old_class_rows = read_csv(project_root / "company_classification_verified_v1.csv")
    company_rows, summary_rows = build_company_rows(all_contacts, old_class_rows, new_drafts)
    class_by_company = {row["company_normalized"]: row for row in company_rows}

    enriched_contacts: list[dict[str, object]] = []
    for contact in all_contacts:
        company = contact.get("company_normalized") or contact.get("company")
        class_row = class_by_company[company]
        out = dict(contact)
        for column in ENRICHED_APPEND_COLUMNS:
            out[column] = class_row.get(column, "")
        enriched_contacts.append(out)

    write_csv(project_root / "namecard_contacts_template_v1.csv", all_contacts, CONTACT_COLUMNS)
    write_csv(project_root / "namecard_contacts_enriched_v1.csv", enriched_contacts, ENRICHED_COLUMNS)
    write_csv(project_root / "company_classification_v1.csv", company_rows, COMPANY_COLUMNS)
    write_csv(project_root / "company_classification_verified_v1.csv", company_rows, VERIFIED_COMPANY_COLUMNS)

    readme_rows = [
        {"field": "用途", "value": "船舶行业名片联系人模板、公司分类和 CRM 导入前结构化数据。"},
        {"field": "数据状态", "value": "扫描 PDF/OCR 后的抽取结果，review_status=待校对 表示建议后续核对。"},
        {"field": "新增工具", "value": "NamecardUpdater.exe 免费本地工具追加联系人并生成网页数据。"},
        {"field": "去重建议", "value": "优先使用 email_primary 去重；无邮箱时用 person_name + company_normalized。"},
    ]
    lookup_rows = []
    for option in ["船东/船管", "船厂", "班轮公司", "船舶经纪", "港口/码头", "物流代理", "主机/发动机", "推进系统", "LNG/低温/燃气", "泵阀", "安全消防", "节能环保", "船级社/认证", "材料", "导航/自动化", "能源", "非船舶"]:
        lookup_rows.append({"field": "industry_tags", "option": option})
    for option in ["待校对", "已确认", "需补充", "可能重复", "暂不使用"]:
        lookup_rows.append({"field": "review_status", "option": option})
    for option in ["high", "medium", "low"]:
        lookup_rows.append({"field": "confidence", "option": option})

    companies_draft = [
        {
            "company_normalized": row["company_normalized"],
            "contact_count": row["contact_count"],
            "country_samples": row["country_samples"],
            "website_samples": row["website_samples"],
            "industry_tags": row["industry_tags_from_cards"],
            "source_pages": row["source_pages"],
            "review_status": "待校对",
            "notes": "",
        }
        for row in company_rows
    ]

    source_pages: list[dict[str, object]] = []
    seen_pages: set[tuple[str, str]] = set()
    for contact in all_contacts:
        key = (contact.get("source_pdf", ""), contact.get("source_page", ""))
        if not key[0] or key in seen_pages:
            continue
        seen_pages.add(key)
        source_pages.append(
            {
                "source_pdf": key[0],
                "source_page": key[1],
                "source_image": contact.get("source_image", ""),
                "notes": "source page",
            }
        )

    write_workbook(
        project_root / "namecard_contacts_template_v1.xlsx",
        [
            ("contacts", CONTACT_COLUMNS, all_contacts),
            ("README", ["field", "value"], readme_rows),
            ("lookups", ["field", "option"], lookup_rows),
            ("companies_draft", ["company_normalized", "contact_count", "country_samples", "website_samples", "industry_tags", "source_pages", "review_status", "notes"], companies_draft),
            ("source_pages", ["source_pdf", "source_page", "source_image", "notes"], source_pages),
        ],
    )
    write_workbook(project_root / "namecard_contacts_enriched_v1.xlsx", [("contacts", ENRICHED_COLUMNS, enriched_contacts)])
    write_workbook(
        project_root / "company_classification_v1.xlsx",
        [
            ("company_classification", COMPANY_COLUMNS, company_rows),
            ("category_summary", ["category_code", "primary_category", "company_count", "contact_count", "definition"], summary_rows),
            ("README", ["field", "value"], readme_rows),
        ],
    )
    write_workbook(
        project_root / "company_classification_verified_v1.xlsx",
        [
            ("company_classification", VERIFIED_COMPANY_COLUMNS, company_rows),
            ("category_summary", ["category_code", "primary_category", "company_count", "contact_count", "definition"], summary_rows),
        ],
    )

    version = generate_mobile_json(project_root, enriched_contacts, company_rows)
    return len(enriched_contacts), len(company_rows), version


def generate_mobile_json(project_root: Path, enriched_contacts: list[dict[str, object]], company_rows: list[dict[str, object]]) -> str:
    output_dir = project_root / "mobile_data"
    web_dir = project_root / "mobile_search"
    output_dir.mkdir(parents=True, exist_ok=True)
    web_dir.mkdir(parents=True, exist_ok=True)
    source_csv = project_root / "namecard_contacts_enriched_v1.csv"
    digest = hashlib.sha256(source_csv.read_bytes()).hexdigest().lower()[:12]
    contacts = []
    for row in enriched_contacts:
        contacts.append({field_name: str(row.get(field_name, "") or "") for field_name in WEB_FIELDS})
    payload = {
        "meta": {
            "generatedAt": datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %z"),
            "sourceCsv": source_csv.name,
            "sourceLastWriteTime": datetime.fromtimestamp(source_csv.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "dataVersion": digest,
            "contactCount": len(contacts),
            "companyCount": len(company_rows),
        },
        "contacts": contacts,
    }
    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    (output_dir / "namecard_contacts_data.json").write_text(text, encoding="utf-8")
    (web_dir / "contacts-data.json").write_text(text, encoding="utf-8")

    package_path = project_root / "mobile_search_update.zip"
    package_files = ["index.html", "styles.css", "app-online-sync-v2.js", "sw.js", "manifest.webmanifest", "icon.svg", "contacts-data.json"]
    with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_name in package_files:
            file_path = web_dir / file_name
            if file_path.is_file():
                archive.write(file_path, arcname=file_name)
    return digest


def git_publish(project_root: Path, message: str) -> str:
    git = shutil.which("git")
    if not git:
        raise RuntimeError("没有找到 git。请先安装 Git，或在 PATH 中加入 git。")
    subprocess.run([git, "add", "--", "mobile_search/contacts-data.json"], cwd=project_root, check=True, capture_output=True, text=True)
    diff = subprocess.run([git, "diff", "--cached", "--quiet"], cwd=project_root)
    if diff.returncode == 0:
        return "没有检测到网页数据变化，跳过提交。"
    commit = subprocess.run([git, "commit", "-m", message], cwd=project_root, capture_output=True, text=True)
    if commit.returncode != 0:
        raise RuntimeError(commit.stderr or commit.stdout)
    push = subprocess.run([git, "push"], cwd=project_root, capture_output=True, text=True)
    if push.returncode != 0:
        raise RuntimeError(push.stderr or push.stdout)
    return (commit.stdout + "\n" + push.stdout + "\n" + push.stderr).strip()


def check_online(expected_version: str) -> str:
    url = f"{PAGES_URL}contacts-data.json?check={int(time.time())}"
    try:
        with urllib.request.urlopen(url, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        return f"线上检查失败：{exc}"
    meta = payload.get("meta", {})
    version = meta.get("dataVersion", "")
    contacts = meta.get("contactCount", "")
    companies = meta.get("companyCount", "")
    matched = "已更新" if version == expected_version else "版本仍在等待 Pages 刷新"
    return f"{matched}：contacts={contacts}, companies={companies}, version={version}"


class NamecardUpdaterApp:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("NamecardUpdater 免费名片新增工具")
        self.root.geometry("1280x820")
        self.project_dir = StringVar(value=str(find_project_root()))
        self.pdf_path = StringVar(value="")
        self.ocr_status = StringVar(value="")
        self.ocr_language = StringVar(value="")
        self.replace_same_pdf = BooleanVar(value=True)
        self.publish_after_write = BooleanVar(value=True)
        self.contacts: list[ContactDraft] = []
        self.current_index: int | None = None
        self.preview_photo = None
        self.editor_vars: dict[str, StringVar] = {}
        self.tesseract_cmd = find_tesseract()
        if self.tesseract_cmd:
            self.ocr_language.set(choose_ocr_language(self.tesseract_cmd))
        else:
            self.ocr_language.set("eng")
        self.build_ui()
        self.refresh_ocr_status()

    def build_ui(self) -> None:
        top = ttk.Frame(self.root, padding=10)
        top.pack(side="top", fill="x")

        ttk.Label(top, text="项目目录").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.project_dir, width=70).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(top, text="选择", command=self.choose_project).grid(row=0, column=2, padx=3)
        ttk.Button(top, text="打开项目", command=lambda: os.startfile(self.project_dir.get())).grid(row=0, column=3, padx=3)

        ttk.Label(top, text="新增 PDF").grid(row=1, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.pdf_path, width=70).grid(row=1, column=1, sticky="ew", padx=5)
        ttk.Button(top, text="选择 PDF", command=self.choose_pdf).grid(row=1, column=2, padx=3)
        ttk.Button(top, text="识别 PDF", command=self.process_pdf_thread).grid(row=1, column=3, padx=3)

        ttk.Label(top, textvariable=self.ocr_status).grid(row=2, column=0, columnspan=2, sticky="w", pady=(6, 0))
        ttk.Label(top, text="OCR语言").grid(row=2, column=2, sticky="e")
        ttk.Entry(top, textvariable=self.ocr_language, width=18).grid(row=2, column=3, sticky="w")
        ttk.Button(top, text="检测 OCR", command=self.refresh_ocr_status).grid(row=2, column=4, padx=3)
        ttk.Button(top, text="安装免费 OCR", command=self.install_ocr).grid(row=2, column=5, padx=3)
        top.columnconfigure(1, weight=1)

        main = ttk.PanedWindow(self.root, orient="horizontal")
        main.pack(fill="both", expand=True, padx=10, pady=6)

        left = ttk.Frame(main)
        right = ttk.Frame(main)
        main.add(left, weight=3)
        main.add(right, weight=2)

        columns = ("record_id", "person_name", "company", "title", "email", "mobile", "category", "confidence", "duplicate")
        self.tree = ttk.Treeview(left, columns=columns, show="headings", height=18)
        headings = {
            "record_id": "编号",
            "person_name": "姓名",
            "company": "公司",
            "title": "职位",
            "email": "邮箱",
            "mobile": "手机",
            "category": "分类",
            "confidence": "置信度",
            "duplicate": "重复检查",
        }
        widths = {"record_id": 70, "person_name": 140, "company": 210, "title": 220, "email": 210, "mobile": 130, "category": 70, "confidence": 80, "duplicate": 150}
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.tag_configure("duplicate", foreground="#B42318")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        buttons = ttk.Frame(left)
        buttons.pack(fill="x", pady=6)
        ttk.Button(buttons, text="新增空行", command=self.add_blank).pack(side="left", padx=3)
        ttk.Button(buttons, text="删除选中", command=self.delete_selected).pack(side="left", padx=3)
        ttk.Button(buttons, text="保存当前行", command=self.save_current).pack(side="left", padx=3)
        ttk.Button(buttons, text="检查重复", command=self.check_duplicates_ui).pack(side="left", padx=3)
        ttk.Checkbutton(buttons, text="替换同 PDF 既有记录", variable=self.replace_same_pdf).pack(side="left", padx=12)
        ttk.Checkbutton(buttons, text="写入后发布", variable=self.publish_after_write).pack(side="left", padx=6)
        ttk.Button(buttons, text="写入并发布", command=self.write_publish_thread).pack(side="right", padx=3)

        form = ttk.LabelFrame(right, text="联系人确认", padding=8)
        form.pack(fill="x")
        fields = [
            ("person_name", "姓名"),
            ("chinese_name", "中文名"),
            ("company", "公司名"),
            ("company_normalized", "标准公司"),
            ("title", "职位"),
            ("country", "国家"),
            ("city", "城市"),
            ("email_primary", "邮箱"),
            ("mobile", "手机"),
            ("phone", "电话"),
            ("website", "官网"),
            ("address", "地址"),
            ("industry_tags", "行业标签"),
            ("notes", "备注"),
            ("primary_category_code", "分类代码"),
            ("subcategory", "子分类"),
            ("business_role", "业务角色"),
            ("classification_confidence", "分类置信度"),
        ]
        for row_index, (key, label) in enumerate(fields):
            ttk.Label(form, text=label).grid(row=row_index, column=0, sticky="w", pady=2)
            var = StringVar()
            self.editor_vars[key] = var
            if key == "primary_category_code":
                widget = ttk.Combobox(form, textvariable=var, values=list(CATEGORIES.keys()), width=42)
            elif key == "classification_confidence":
                widget = ttk.Combobox(form, textvariable=var, values=["high", "medium", "low"], width=42)
            else:
                widget = ttk.Entry(form, textvariable=var, width=46)
            widget.grid(row=row_index, column=1, sticky="ew", pady=2)
        form.columnconfigure(1, weight=1)

        preview_box = ttk.LabelFrame(right, text="名片图像 / OCR文本", padding=8)
        preview_box.pack(fill="both", expand=True, pady=6)
        self.preview_label = ttk.Label(preview_box)
        self.preview_label.pack(fill="x")
        self.ocr_text = ttk.Treeview(preview_box, columns=("text",), show="headings", height=5)
        self.ocr_text.heading("text", text="OCR文本")
        self.ocr_text.column("text", width=420, anchor="w")
        self.ocr_text.pack(fill="both", expand=True, pady=(6, 0))

        log_box = ttk.LabelFrame(self.root, text="运行日志", padding=6)
        log_box.pack(fill="x", padx=10, pady=(0, 10))
        self.log_view = ttk.Treeview(log_box, columns=("log",), show="headings", height=5)
        self.log_view.heading("log", text="日志")
        self.log_view.column("log", width=1180, anchor="w")
        self.log_view.pack(fill="x")

    def log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_view.insert("", "end", values=(f"[{timestamp}] {message}",))
        self.log_view.yview_moveto(1.0)
        self.root.update_idletasks()

    def choose_project(self) -> None:
        path = filedialog.askdirectory(initialdir=self.project_dir.get() or "E:/")
        if path:
            self.project_dir.set(path)

    def choose_pdf(self) -> None:
        path = filedialog.askopenfilename(initialdir=self.project_dir.get() or "E:/", filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if path:
            self.pdf_path.set(path)

    def refresh_ocr_status(self) -> None:
        self.tesseract_cmd = find_tesseract()
        if self.tesseract_cmd:
            langs = tesseract_languages(self.tesseract_cmd)
            if not self.ocr_language.get() or self.ocr_language.get() == "eng":
                self.ocr_language.set(choose_ocr_language(self.tesseract_cmd))
            self.ocr_status.set(f"OCR 已找到：{self.tesseract_cmd}；语言：{', '.join(langs[:8])}")
        else:
            self.ocr_status.set("未找到 OCR。仍可手工录入；可点击“安装免费 OCR”。")

    def install_ocr(self) -> None:
        winget = shutil.which("winget")
        if not winget:
            messagebox.showinfo("安装 OCR", "未找到 winget。请手动安装 Tesseract OCR：UB Mannheim Tesseract。")
            webbrowser.open("https://github.com/UB-Mannheim/tesseract/wiki")
            return
        if not messagebox.askyesno("安装免费 OCR", "将通过 winget 安装 UB-Mannheim Tesseract OCR，可能弹出系统确认窗口。继续吗？"):
            return
        subprocess.Popen([winget, "install", "--id", "UB-Mannheim.TesseractOCR", "-e"])
        self.log("已启动 winget 安装。安装完成后点击“检测 OCR”。")

    def add_blank(self) -> None:
        pdf = Path(self.pdf_path.get()) if self.pdf_path.get() else Path("")
        data = {column: "" for column in ENRICHED_COLUMNS}
        data.update(
            {
                "review_status": "待校对",
                "confidence": "medium",
                "source_pdf": str(pdf) if pdf else "",
                "source_page": "1",
                "card_position": "手工新增",
                "relationship_stage": "名片已扫描",
                "primary_category_code": "C09",
                "primary_category": CATEGORIES["C09"][0],
                "classification_confidence": "medium",
            }
        )
        self.contacts.append(ContactDraft(data=data))
        self.refresh_table(select_index=len(self.contacts) - 1)

    def delete_selected(self) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        indexes = sorted([int(item) for item in selection], reverse=True)
        for index in indexes:
            if 0 <= index < len(self.contacts):
                del self.contacts[index]
        self.current_index = None
        self.refresh_table()

    def on_select(self, _event=None) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        index = int(selection[0])
        if 0 <= index < len(self.contacts):
            self.current_index = index
            self.load_editor(self.contacts[index])

    def load_editor(self, draft: ContactDraft) -> None:
        for key, var in self.editor_vars.items():
            var.set(draft.get(key))
        self.show_preview(draft.crop_image or draft.get("source_image"))
        self.ocr_text.delete(*self.ocr_text.get_children())
        for line in draft.ocr_text.splitlines()[:80]:
            self.ocr_text.insert("", "end", values=(line,))

    def show_preview(self, path: str) -> None:
        if Image is None or ImageTk is None or not path or not Path(path).is_file():
            self.preview_label.configure(image="", text="无图像预览")
            self.preview_photo = None
            return
        image = Image.open(path)
        image.thumbnail((420, 260))
        self.preview_photo = ImageTk.PhotoImage(image)
        self.preview_label.configure(image=self.preview_photo, text="")

    def save_current(self) -> None:
        if self.current_index is None or not (0 <= self.current_index < len(self.contacts)):
            return
        draft = self.contacts[self.current_index]
        for key, var in self.editor_vars.items():
            draft.set(key, var.get())
        code = draft.get("primary_category_code")
        if code in CATEGORIES:
            draft.set("primary_category", CATEGORIES[code][0])
        if not draft.get("company_normalized"):
            draft.set("company_normalized", draft.get("company"))
        if not draft.get("industry_tags"):
            draft.set("industry_tags", infer_tags(draft.get("company_normalized"), draft.get("title"), draft.ocr_text))
        self.refresh_table(select_index=self.current_index)

    def duplicate_reports(self) -> dict[int, list[str]]:
        try:
            project = Path(self.project_dir.get()).resolve()
            existing = load_template_contacts(project)
        except Exception:
            existing = []
        return find_duplicate_reports(self.contacts, existing, self.replace_same_pdf.get())

    def duplicate_report_text(self, reports: dict[int, list[str]]) -> str:
        lines: list[str] = []
        for index in sorted(reports):
            draft = self.contacts[index]
            title = contact_label(draft.data, fallback=f"第 {index + 1} 条")
            lines.append(f"第 {index + 1} 条：{title}")
            for item in reports[index][:6]:
                lines.append(f"  - {item}")
        return "\n".join(lines)

    def check_duplicates_ui(self) -> None:
        if self.current_index is not None:
            self.save_current()
        reports = self.duplicate_reports()
        if not reports:
            messagebox.showinfo("重复检查", "没有发现明显重复。")
            self.log("重复检查：没有发现明显重复。")
            return
        text = self.duplicate_report_text(reports)
        if len(text) > 3500:
            text = text[:3500] + "\n..."
        messagebox.showwarning("发现可能重复", text)
        self.log(f"重复检查：发现 {len(reports)} 条候选联系人可能重复。")
        self.refresh_table(select_index=self.current_index)

    def refresh_table(self, select_index: int | None = None) -> None:
        self.tree.delete(*self.tree.get_children())
        reports = self.duplicate_reports() if self.contacts else {}
        for index, draft in enumerate(self.contacts):
            duplicate_text = duplicate_summary(reports.get(index, []))
            self.tree.insert(
                "",
                "end",
                iid=str(index),
                tags=("duplicate",) if duplicate_text else (),
                values=(
                    draft.get("record_id"),
                    draft.get("person_name") or draft.get("chinese_name"),
                    draft.get("company_normalized") or draft.get("company"),
                    draft.get("title"),
                    draft.get("email_primary"),
                    draft.get("mobile"),
                    draft.get("primary_category_code"),
                    draft.get("confidence"),
                    duplicate_text,
                ),
            )
        if select_index is not None and 0 <= select_index < len(self.contacts):
            self.tree.selection_set(str(select_index))
            self.tree.focus(str(select_index))
            self.current_index = select_index
            self.load_editor(self.contacts[select_index])

    def process_pdf_thread(self) -> None:
        threading.Thread(target=self.process_pdf, daemon=True).start()

    def process_pdf(self) -> None:
        try:
            project = Path(self.project_dir.get()).resolve()
            pdf = Path(self.pdf_path.get()).resolve()
            if not pdf.is_file():
                raise RuntimeError("请选择有效 PDF。")
            existing = load_template_contacts(project)
            out_dir = project / "_analysis" / "updater" / pdf.stem
            self.log(f"渲染 PDF：{pdf.name}")
            pages = render_pdf(pdf, out_dir / "pages")
            self.log(f"已渲染 {len(pages)} 页，开始切分名片。")
            tesseract_cmd = self.tesseract_cmd
            lang = self.ocr_language.get() or "eng"
            drafts: list[ContactDraft] = []
            for page_index, page_image in enumerate(pages, start=1):
                crops = detect_card_crops(page_image, out_dir / "crops")
                self.log(f"第 {page_index} 页候选名片：{len(crops)}")
                for position, crop_path in crops:
                    text = run_ocr(tesseract_cmd, crop_path, lang) if tesseract_cmd else ""
                    draft = parse_contact_from_ocr(text, pdf, page_index, position, crop_path, existing)
                    if not text:
                        draft.ocr_text = "未运行 OCR。请手工填写联系人字段。"
                    drafts.append(draft)
            self.contacts = drafts
            self.root.after(0, lambda: self.refresh_table(select_index=0 if self.contacts else None))
            reports = find_duplicate_reports(self.contacts, existing, self.replace_same_pdf.get())
            if reports:
                self.log(f"识别完成：{len(self.contacts)} 条候选联系人，其中 {len(reports)} 条可能重复。请检查“重复检查”列。")
            else:
                self.log(f"识别完成：{len(self.contacts)} 条候选联系人，未发现明显重复。")
        except Exception as exc:
            self.log(f"失败：{exc}")
            messagebox.showerror("识别失败", str(exc))

    def write_publish_thread(self) -> None:
        if self.current_index is not None:
            self.save_current()
        threading.Thread(target=self.write_publish, daemon=True).start()

    def write_publish(self) -> None:
        try:
            project = Path(self.project_dir.get()).resolve()
            for index, draft in enumerate(self.contacts, start=1):
                if not draft.get("person_name") and not draft.get("chinese_name"):
                    raise RuntimeError(f"第 {index} 行缺少姓名。")
                if not draft.get("company_normalized") and not draft.get("company"):
                    raise RuntimeError(f"第 {index} 行缺少公司。")
            reports = self.duplicate_reports()
            if reports:
                text = self.duplicate_report_text(reports)
                if len(text) > 3500:
                    text = text[:3500] + "\n..."
                proceed = messagebox.askyesno(
                    "发现可能重复",
                    f"发现 {len(reports)} 条候选联系人可能重复。\n\n{text}\n\n仍然写入并继续发布吗？",
                )
                if not proceed:
                    self.log("已取消写入：存在可能重复的联系人。")
                    return
            self.log("写回主表并生成网页数据。")
            contacts, companies, version = regenerate_outputs(project, self.contacts, self.replace_same_pdf.get())
            self.log(f"已生成：contacts={contacts}, companies={companies}, version={version}")
            if self.publish_after_write.get():
                pdf_name = Path(self.pdf_path.get()).name or "new PDF"
                self.log("提交并推送 GitHub。")
                result = git_publish(project, f"Add contacts from {pdf_name}")
                self.log(result or "Git 推送完成。")
                self.log("等待 GitHub Pages 刷新。")
                time.sleep(25)
                self.log(check_online(version))
            messagebox.showinfo("完成", f"已完成。\n联系人：{contacts}\n公司：{companies}\n版本：{version}")
        except Exception as exc:
            self.log(f"失败：{exc}")
            messagebox.showerror("写入/发布失败", str(exc))


def command_check() -> int:
    root = find_project_root()
    rows = read_csv(root / "namecard_contacts_enriched_v1.csv")
    tesseract_cmd = find_tesseract()
    print(f"project={root}")
    print(f"contacts={len(rows)}")
    print(f"tesseract={tesseract_cmd or 'not_found'}")
    if tesseract_cmd:
        print(f"languages={','.join(tesseract_languages(tesseract_cmd))}")
    try:
        import pypdfium2  # noqa: F401
        print("pypdfium2=ok")
    except Exception as exc:
        print(f"pypdfium2=missing:{exc}")
        return 1
    try:
        import openpyxl  # noqa: F401
        print("openpyxl=ok")
    except Exception as exc:
        print(f"openpyxl=missing:{exc}")
        return 1
    if Image is None:
        print("pillow=missing")
        return 1
    print("pillow=ok")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="NamecardUpdater")
    parser.add_argument("--check", action="store_true", help="check dependencies and project")
    args = parser.parse_args()
    if args.check:
        return command_check()
    root = Tk()
    app = NamecardUpdaterApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
